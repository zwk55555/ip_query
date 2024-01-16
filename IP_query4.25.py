import requests
import openpyxl
import gc

#query_ip = []
my_key = "your_key"
urls = "https://api.ipplus360.com/ip/info/v1/scene/?key=" + my_key + "&ip="

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/99.0.4844.51 Safari/537.36'}


# 获取Excel表中的IP数据并返回
def get_ip_data():
    # data = []

    if IP_number == "" or type(IP_number) != int:
        print("请输入正确的IP数量")
        quit()
    wb = openpyxl.load_workbook('ip_query.xlsx')
    sheet_ip = wb['Sheet1']

    for i in range(1, IP_number + 1):
        ip = str(sheet_ip.cell(row=i, column=1).value)
        ip = ip.strip()
        api_url = urls + ip
        rsp = requests.get(api_url, headers=headers)
        #print(rsp.json())
        sheet_ip.cell(row=i, column=2).value = rsp.json()['data']['scene']
        print("ip     " + str(sheet_ip.cell(row=i, column=1).value) + "------" + "应用场景为：" + rsp.json()['data']['scene'])

    wb.save('ip_query.xlsx')

    del wb
    gc.collect()

if __name__ == '__main__':
    IP_number = int(input("请输入需要查询的IP数量："))
    get_ip_data()
